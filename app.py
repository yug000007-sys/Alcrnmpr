import io
import re
import zipfile
import streamlit as st
import pandas as pd
import pdfplumber

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# -----------------------------
# AUTH
# -----------------------------
USERNAME = "matt"
PASSWORD = "Interlynx123"

if "auth" not in st.session_state:
    st.session_state.auth = False

if not st.session_state.auth:
    st.title("Alcorn PDF Extractor — Login")
    u = st.text_input("Username")
    p = st.text_input("Password", type="password")
    if st.button("Login"):
        if u == USERNAME and p == PASSWORD:
            st.session_state.auth = True
            st.rerun()
        else:
            st.error("Invalid username or password.")
    st.stop()

st.title("Alcorn PDF → Excel (Auto-Formatted)")

# -----------------------------
# EXACT HEADER (YOUR REQUIRED)
# -----------------------------
OUTPUT_COLUMNS = [
    "ReferralManagerCode","ReferralManager","ReferralEmail","Brand","QuoteNumber","QuoteVersion",
    "QuoteDate","QuoteValidDate","Customer Number/ID","Company","Address","County","City","State",
    "ZipCode","Country","FirstName","LastName","ContactEmail","ContactPhone","Webaddress","item_id",
    "item_desc","UOM","Quantity","Unit Price","List Price","TotalSales","Manufacturer_ID",
    "manufacturer_Name","Writer Name","CustomerPONumber","PDF","DemoQuote","Duns","SIC","NAICS",
    "LineOfBusiness","LinkedinProfile","PhoneResearched","PhoneSupplied","ParentName"
]

US_STATE_RE = r"(AL|AK|AZ|AR|CA|CO|CT|DE|FL|GA|HI|IA|ID|IL|IN|KS|KY|LA|MA|MD|ME|MI|MN|MO|MS|MT|NC|ND|NE|NH|NJ|NM|NV|NY|OH|OK|OR|PA|RI|SC|SD|TN|TX|UT|VA|VT|WA|WI|WV|WY|DC)"
CAN_POSTAL_RE = r"\b([A-Z]\d[A-Z]\s?\d[A-Z]\d)\b"
MONEY_RE = re.compile(r"\d{1,3}(?:,\d{3})*\.\d{2}")

def make_row():
    return {c: "" for c in OUTPUT_COLUMNS}

def norm(x):
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    return str(x).strip().replace("\u00a0"," ")

def quote_norm(x):
    return norm(x).replace("'", "").replace('"', "").upper()

def to_mmddyyyy(val):
    s = norm(val)
    if not s:
        return ""
    dt = pd.to_datetime(s, errors="coerce", infer_datetime_format=True)
    if pd.notna(dt):
        return dt.strftime("%m/%d/%Y")
    m = re.search(r"\b([A-Za-z]{3,9})\s+(\d{1,2}),\s*(\d{4})\b", s)
    if m:
        mon = m.group(1).lower()[:3]
        mon_map = {"jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,"jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}
        if mon in mon_map:
            return f"{mon_map[mon]:02d}/{int(m.group(2)):02d}/{int(m.group(3))}"
    return s

def money_float(s):
    t = norm(s).replace(",", "")
    try:
        return float(t)
    except:
        return 0.0

def extract_full_text(pdf):
    parts = []
    for p in pdf.pages:
        t = p.extract_text() or ""
        if t.strip():
            parts.append(t)
    return "\n".join(parts)

def extract_header_fields(full_text: str):
    out = {
        "ReferralManagerCode": "",
        "QuoteNumber": "",
        "QuoteDate": "",
        "Customer Number/ID": "",
        "Brand": "Alcorn Industrial Inc",
    }

    m = re.search(r"\bOrder Number\s*(QT[0-9A-Z]+)\b", full_text, flags=re.IGNORECASE)
    if m:
        out["QuoteNumber"] = quote_norm(m.group(1))
    else:
        m2 = re.search(r"\b(QT[0-9A-Z]+)\b", full_text, flags=re.IGNORECASE)
        out["QuoteNumber"] = quote_norm(m2.group(1)) if m2 else ""

    md = re.search(r"\bDate\s*([A-Za-z]{3,9}\s+\d{1,2},\s*\d{4})\b", full_text, flags=re.IGNORECASE)
    out["QuoteDate"] = to_mmddyyyy(md.group(1)) if md else ""

    mc = re.search(r"\bCustomer No\.\s*([0-9A-Z\-]+)\b", full_text, flags=re.IGNORECASE)
    out["Customer Number/ID"] = norm(mc.group(1)) if mc else ""

    ms = re.search(r"\bSalesperson\s*([A-Z]{1,4})\b", full_text, flags=re.IGNORECASE)
    out["ReferralManagerCode"] = norm(ms.group(1)) if ms else ""

    return out

def extract_ship_to(full_text: str):
    out = {"Company":"", "Address":"", "City":"", "State":"", "ZipCode":"", "Country":""}

    # Grab block after "Ship To:" up to the next boxed header area (Customer No./Salesperson)
    m = re.search(r"Ship To:\s*(.+?)(?:\n\s*Reference|\n\s*PO Number|\n\s*Customer No\.|\n\s*Salesperson)",
                  full_text, flags=re.IGNORECASE | re.DOTALL)
    if not m:
        return out

    block = m.group(1).strip()
    lines = [l.strip() for l in block.splitlines() if l.strip()]
    if not lines:
        return out

    out["Company"] = lines[0]

    street, cityline, country = "", "", ""
    for l in lines[1:]:
        if not street and re.search(r"\d", l):
            street = l
            continue
        if not cityline and ("," in l) and (re.search(r"\b[A-Z]{2}\b", l) or re.search(CAN_POSTAL_RE, l)):
            cityline = l
            continue
        if l.lower() in ("canada", "usa", "united states", "united states of america"):
            country = "Canada" if "canada" in l.lower() else "USA"

    out["Address"] = street
    out["Country"] = country

    if cityline:
        # Canada: City, QC J7E4K9
        mca = re.search(r"^(.*?),\s*([A-Z]{2})\s*,?\s*([A-Z]\d[A-Z]\s?\d[A-Z]\d)\s*$", cityline)
        if mca:
            out["City"] = mca.group(1).strip()
            out["State"] = mca.group(2).strip()
            out["ZipCode"] = mca.group(3).replace(" ", "").strip()
        else:
            mus = re.search(rf"^(.*?),\s*({US_STATE_RE})\s*,?\s*(\d{{5}}(?:-\d{{4}})?)\s*$", cityline)
            if mus:
                out["City"] = mus.group(1).strip()
                out["State"] = mus.group(2).strip()
                out["ZipCode"] = mus.group(3).strip()

    return out

def group_lines(words, y_tol=3):
    """Group extracted words into visual lines by y position."""
    lines = []
    for w in words:
        placed = False
        for ln in lines:
            if abs(ln["top"] - w["top"]) <= y_tol:
                ln["words"].append(w)
                placed = True
                break
        if not placed:
            lines.append({"top": w["top"], "words":[w]})
    for ln in lines:
        ln["words"] = sorted(ln["words"], key=lambda x: x["x0"])
    lines = sorted(lines, key=lambda x: x["top"])
    return lines

def find_header_line(lines):
    """Find the table header line containing key labels."""
    header_keys = ["Qty.", "Customer", "Description", "Unit", "Extended"]
    for ln in lines:
        text = " ".join([w["text"] for w in ln["words"]]).lower()
        hit = sum(1 for k in header_keys if k.lower() in text)
        if hit >= 3:
            return ln
    return None

def detect_columns_from_header(header_ln):
    """
    Detect approximate x boundaries from header words.
    We want these columns:
      Qty. Ord. | Customer Item Number | Description | Unit Price | Extended Price
    """
    # Get x positions of anchor words
    anchors = []
    for w in header_ln["words"]:
        t = w["text"].lower()
        if "qty" in t:
            anchors.append(("qty", w["x0"]))
        if "customer" in t:
            anchors.append(("cust", w["x0"]))
        if "description" in t:
            anchors.append(("desc", w["x0"]))
        if "unit" in t:
            anchors.append(("unit", w["x0"]))
        if "extended" in t:
            anchors.append(("ext", w["x0"]))

    # Fallback fixed positions if detection fails
    x_qty = min([x for k,x in anchors if k=="qty"], default=0)
    x_cust = min([x for k,x in anchors if k=="cust"], default=150)
    x_desc = min([x for k,x in anchors if k=="desc"], default=280)
    x_unit = min([x for k,x in anchors if k=="unit"], default=520)
    x_ext  = min([x for k,x in anchors if k=="ext"],  default=620)

    # Build ranges
    col_qty = (x_qty - 5, x_cust - 10)
    col_cust = (x_cust - 5, x_desc - 10)
    col_desc = (x_desc - 5, x_unit - 10)
    col_unit = (x_unit - 5, x_ext - 10)
    col_ext  = (x_ext - 5, 9999)

    return col_qty, col_cust, col_desc, col_unit, col_ext

def line_text_in_col(words, col):
    return " ".join([w["text"] for w in words if col[0] <= w["x0"] < col[1]]).strip()

def extract_items_from_page(page: pdfplumber.page.Page):
    """
    Robust extraction of items using word positions.
    item_id = Customer Item Number (your light green)
    item_desc = Description (pink box)
    Quantity = Qty. Ord.
    Unit Price and TotalSales = monetary columns
    """
    words = page.extract_words(x_tolerance=2, y_tolerance=2, keep_blank_chars=False)
    if not words:
        return []

    lines = group_lines(words, y_tol=3)
    header_ln = find_header_line(lines)
    if not header_ln:
        return []

    col_qty, col_cust, col_desc, col_unit, col_ext = detect_columns_from_header(header_ln)
    header_top = header_ln["top"]

    data_lines = [ln for ln in lines if ln["top"] > header_top + 5]

    items = []
    current = None

    for ln in data_lines:
        ws = ln["words"]

        qty_txt = line_text_in_col(ws, col_qty)

        # New row if qty is numeric
        if re.fullmatch(r"\d+", qty_txt):
            if current:
                current["item_desc"] = current["item_desc"].strip()
                items.append(current)

            cust_txt = line_text_in_col(ws, col_cust)
            desc_txt = line_text_in_col(ws, col_desc)
            unit_txt = line_text_in_col(ws, col_unit)
            ext_txt  = line_text_in_col(ws, col_ext)

            # find last money values
            unit_m = MONEY_RE.findall(unit_txt)
            ext_m  = MONEY_RE.findall(ext_txt)

            unit_price = unit_m[-1] if unit_m else ""
            ext_price  = ext_m[-1] if ext_m else ""

            current = {
                "Quantity": qty_txt,
                "item_id": cust_txt,
                "item_desc": desc_txt,
                "Unit Price": money_float(unit_price),
                "TotalSales": money_float(ext_price),
            }
        else:
            # Continuation line -> append description
            if current:
                wrap = line_text_in_col(ws, col_desc)
                if wrap:
                    current["item_desc"] += " " + wrap

    if current:
        current["item_desc"] = current["item_desc"].strip()
        items.append(current)

    # Remove empty
    items = [it for it in items if it.get("item_id") or it.get("item_desc")]
    return items

def auto_format_excel(xlsx_bytes: bytes) -> bytes:
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    TEXT_COLS = {
        "ReferralManagerCode","QuoteNumber","QuoteVersion","Customer Number/ID",
        "ZipCode","item_id","Manufacturer_ID","Duns","SIC","NAICS","CustomerPONumber","PDF"
    }
    DATE_COLS = {"QuoteDate","QuoteValidDate"}
    CURR_COLS = {"Unit Price","List Price","TotalSales"}
    NUM_COLS = {"Quantity"}

    header_map = {str(c.value).strip(): i for i, c in enumerate(ws[1], start=1)}
    max_row = ws.max_row

    for col_name, j in header_map.items():
        col_letter = get_column_letter(j)

        if col_name in TEXT_COLS:
            for r in range(2, max_row + 1):
                ws[f"{col_letter}{r}"].number_format = "@"

        if col_name in DATE_COLS:
            for r in range(2, max_row + 1):
                ws[f"{col_letter}{r}"].number_format = "mm/dd/yyyy"

        if col_name in CURR_COLS:
            for r in range(2, max_row + 1):
                ws[f"{col_letter}{r}"].number_format = '"$"#,##0.00'

        if col_name in NUM_COLS:
            for r in range(2, max_row + 1):
                ws[f"{col_letter}{r}"].number_format = "0"

        ws.column_dimensions[col_letter].width = min(45, max(12, len(col_name) + 2))

    if "item_desc" in header_map:
        ws.column_dimensions[get_column_letter(header_map["item_desc"])].width = 55
    if "Company" in header_map:
        ws.column_dimensions[get_column_letter(header_map["Company"])].width = 35
    if "Address" in header_map:
        ws.column_dimensions[get_column_letter(header_map["Address"])].width = 35

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()

# ---------------- UI ----------------
pdfs = st.file_uploader("Upload Alcorn Quote PDFs (up to 100)", type=["pdf"], accept_multiple_files=True)

if st.button("Extract + Auto-Format Excel"):
    if not pdfs:
        st.error("Please upload at least 1 PDF.")
        st.stop()

    rows = []
    renamed = []

    for f in pdfs:
        pdf_bytes = f.read()
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            full_text = extract_full_text(pdf)

            header = extract_header_fields(full_text)
            ship = extract_ship_to(full_text)

            quote = quote_norm(header.get("QuoteNumber") or f.name.replace(".pdf",""))
            pdf_name = f"{quote}.pdf"
            renamed.append((pdf_name, pdf_bytes))

            # Extract line items from page 1 (most Alcorn quotes have table on first page)
            items = extract_items_from_page(pdf.pages[0])

            # If still nothing, try page 2 just in case
            if not items and len(pdf.pages) > 1:
                items = extract_items_from_page(pdf.pages[1])

            for it in items:
                r = make_row()
                r["Brand"] = header.get("Brand", "Alcorn Industrial Inc")
                r["QuoteNumber"] = quote
                r["QuoteDate"] = header.get("QuoteDate", "")
                r["Customer Number/ID"] = header.get("Customer Number/ID", "")
                r["ReferralManagerCode"] = header.get("ReferralManagerCode", "")

                r["Company"] = ship.get("Company", "")
                r["Address"] = ship.get("Address", "")
                r["City"] = ship.get("City", "")
                r["State"] = ship.get("State", "")
                r["ZipCode"] = ship.get("ZipCode", "")
                r["Country"] = ship.get("Country", "")

                r["item_id"] = norm(it.get("item_id",""))
                r["item_desc"] = norm(it.get("item_desc",""))
                r["Quantity"] = norm(it.get("Quantity",""))
                r["Unit Price"] = it.get("Unit Price", 0.0)
                r["TotalSales"] = it.get("TotalSales", 0.0)

                r["PDF"] = pdf_name
                rows.append(r)

    df = pd.DataFrame(rows, columns=OUTPUT_COLUMNS)

    st.success(f"Extracted rows: {len(df)}")
    st.dataframe(df.head(25), height=260, use_container_width=True)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Extracted")
    buf.seek(0)

    formatted = auto_format_excel(buf.getvalue())

    st.download_button(
        "Download Auto-Formatted Excel",
        data=formatted,
        file_name="alcorn_extracted_autoformatted.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # Renamed PDFs ZIP
    zbuf = io.BytesIO()
    with zipfile.ZipFile(zbuf, "w", zipfile.ZIP_DEFLATED) as z:
        for name, b in renamed:
            z.writestr(name, b)
    zbuf.seek(0)

    st.download_button(
        "Download Renamed PDFs (ZIP)",
        data=zbuf.getvalue(),
        file_name="alcorn_renamed_pdfs.zip",
        mime="application/zip",
    )
